#test excel_payload_import
import openpyxl

def excel_payload_import():
    workbook = openpyxl.load_workbook("/home/rpl/workspace/BIO_workcell/applications/multi_gc_350_app/example_payload.xlsx")
    sheet = workbook.active
    
    payload = {}

    for row in sheet.iter_rows(min_row=1, max_col=sheet.max_column, max_row=sheet.max_row):
        key = row[0].value
        values = [cell.value for cell in row[1:]]
        values = [x for x in values if x is not None]
        payload[key] = values

    return payload


def populate_payload(payload):
    payload["temp"] = payload["Temperature"]
    del payload["Temperature"]
    payload["temp"] = payload["temp"][0]

    payload["humidity"] = payload["Humidity"]
    del payload["Humidity"]
    payload["humidity"] = payload["humidity"][0]

    payload["shaker_speed"] = payload["Shaker Speed"]
    del payload["Shaker Speed"]
    payload["shaker_speed"] = payload["shaker_speed"][0]

    payload["treatment"] = payload["Treatment Columns"]
    del payload["Treatment Columns"]

    payload["culture_column"] = payload["Culture Columns"]
    del payload["Culture Columns"]

    payload["tip_box_position"] = payload["Tip Box Position"]
    del payload["Tip Box Position"]
    payload["tip_box_position"] = str(payload["tip_box_position"][0])

    # make sure we have same number of treatments and cultures
    if len(payload["treatment"]) != len(payload["culture_column"]):
        payload = "ERROR"
        print("ERROR, DIFFERENT NUMBER OF TREATMENTS AND CULTURES")
    
    payload["media_start_column"] = []
    for i in range(len(payload["treatment"])):
        if i == 0:
            payload["media_start_column"].append(1)
        elif i == 6:
            payload["media_start_column"].append(1)
        else:
            payload["media_start_column"].append(payload["media_start_column"][i-1]+2)


    payload["treatment_dil_half"] = []
    for i in range(len(payload["treatment"])):
        if i == 0:
            payload["treatment_dil_half"].append(1)
        elif payload["treatment_dil_half"][i-1] == 1:
            payload["treatment_dil_half"].append(2)
        elif payload["treatment_dil_half"][i-1] == 2:
            payload["treatment_dil_half"].append(1)

    
    for i in range(len(payload["treatment"])):
        payload["treatment"][i] = "col"+str(payload["treatment"][i])


    return payload

def create_payload():
    payload = excel_payload_import
    payload = populate_payload([payload])
    return payload
