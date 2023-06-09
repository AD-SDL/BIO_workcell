#!/usr/bin/env python3

import logging
from argparse import ArgumentParser
import time
# from tools.c2_flow import c2_flow
from pathlib import Path
# from workflows.hso_functions import package_hso
# from workflows import solo_step1, solo_step2, solo_step3
# from workflows import solo_multi_step1, solo_multi_step2, solo_multi_step3
import pandas as pd 
import pathlib
import openpyxl
import tensorflow as tf
from tensorflow import keras
import numpy as np
import os
import datetime
import random
from openpyxl.utils.dataframe import dataframe_to_rows
from sklearn.model_selection import train_test_split

# from rpl_wei import Experiment

#from rpl_wei.wei_workcell_base import WEI

ORIGINAL_ANTIBIOTIC_CONCENTRATION = [1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
ORIGINAL_CELL_CONCENTRATION = [1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
TOTAL_CELL_COLUMN_CONCENTRATION = [] #Each row represents another plate, each value is the concentration of the cell in the column of the plate
TOTAL_TREATMENT_COLUMN_CONCENTRATION = [] #Each row represents another plate, each value is the concentration of the cell in the column of the plate
CULTURE_PAYLOAD = []
MEDIA_PAYLOAD = []
HIDEX_UPLOADS = []
COMPLETED_CELL_COLUMNS = []
COMPLETED_ANTIBIOTIC_COLUMNS = []
PLATE_BARCODES = []
CREATED_COMPLETED_FILE = False
COMPLETED_FILE_NAME = ''
EXPERIMENT_FILE_PATH = ''

TENSORFLOW_MODEL = None
AI_MODEL_FILE_PATH = str(pathlib.Path().resolve()) + "/tensorflow_model"
AI_MODEL_IN_USE = True

COMPLETE_HUDSON_SETUP_FILE_PATH = '/home/rpl/workspace/BIO_workcell/growth_app/workflows/multiple_growth_curve/complete_hudson_setup.yaml'
STREAMLINED_HUDSON_SETUP_FILE_PATH = '/home/rpl/workspace/BIO_workcell/growth_app/workflows/multiple_growth_curve/streamlined_hudson_setup.yaml'
SETUP_GROWTH_MEDIA_FILE_PATH = '/home/rpl/workspace/BIO_workcell/growth_app/workflows/multiple_growth_curve/setup_growth_media.yaml'

CREATE_PLATE_T0_FILE_PATH = '/home/rpl/workspace/BIO_workcell/growth_app/workflows/multiple_growth_curve/create_plate_T0.yaml'
READ_PLATE_T12_FILE_PATH = '/home/rpl/workspace/BIO_workcell/growth_app/workflows/multiple_growth_curve/read_plate_T12.yaml'

DISPOSE_BOX_PLATE_FILE_PATH = '/home/rpl/workspace/BIO_workcell/growth_app/workflows/multiple_growth_curve/dispose_box_plate.yaml'
DISPOSE_GROWTH_MEDIA_FILE_PATH = '/home/rpl/workspace/BIO_workcell/growth_app/workflows/multiple_growth_curve/dispose_growth_media.yaml'

HIDEX_OPEN_CLOSE_FILE_PATH = '/home/rpl/workspace/BIO_workcell/growth_app/workflows/multiple_growth_curve/open_close_hidex.yaml'

# exp = Experiment('127.0.0.1', '8000', 'Growth_Curve')
# exp.register_exp() 
print("Registered Experiment")
# exp.events.log_local_compute("package_hso")

def main():
    load_model()
    iteration_runs, incubation_time = determine_payload_from_excel()
    run_experiment(iteration_runs, incubation_time)
    process_results()
    train_model()
    save_model()
    #Find a way to calculate possible remaining runs
    for i in range(0, 8):
        load_model()
        predict_experiment(1)
        #Need a way to transfer things here to experiment
        run_experiment(iteration_runs)
        process_results()
        train_model()
        save_model()
    delete_experiment_excel_file()

def load_model():
    global TENSORFLOW_MODEL
    if os.path.exists(AI_MODEL_FILE_PATH):
        TENSORFLOW_MODEL = tf.keras.models.load_model(AI_MODEL_FILE_PATH)
    else:
        input_dim = 4 #Cell type, Antibiotic Type, Concentration 1, Concentration 2
        num_classes = 1 #Growth Rate

        TENSORFLOW_MODEL = keras.Sequential([
            keras.layers.Dense(64, activation='relu', input_shape=(input_dim,)),
            keras.layers.Dense(64, activation='relu'),
            keras.layers.Dense(num_classes, activation='softmax')
        ])

        TENSORFLOW_MODEL.compile(optimizer='adam', loss='mean_squared_error')
        print(TENSORFLOW_MODEL.summary())
    
def predict_experiment(num_prediction_requests):
    global TENSORFLOW_MODEL
    #Predict the Model on the Data frame
    #Sort the combinations and output the one/two/three/...twelve that need AI modeling
    predictions = []
    prediction_df = return_combination_data_frame()
     # Make prediction on the combination using the trained model
    numeric_columns = ['Treatment Column', 'Treatment Concentration', 'Cell Column', 'Cell Concentration']
    prediction_df[numeric_columns] = prediction_df[numeric_columns].apply(pd.to_numeric, errors='coerce')

    prediction = TENSORFLOW_MODEL.predict(prediction_df)
    prediction_df['Predictions'] = prediction

    prediction_df = prediction_df.sort_values('Predictions')

    # Select the set of smallest values defined by num_prediction_requests
    selected_rows = prediction_df.head(num_prediction_requests)

    treatment_column = selected_rows['Treatment Column'].values
    treatment_concentration = selected_rows['Treatment Concentration'].values
    cell_column = selected_rows['Cell Column'].values
    cell_concentration = selected_rows['Cell Concentration'].values
    predictions = selected_rows['Predictions'].values
    print(treatment_column)
    print(treatment_concentration)
    print(cell_column)
    print(cell_concentration)
    print(predictions)

    return treatment_column, treatment_concentration, cell_column, cell_concentration, predictions

def return_combination_data_frame():
    combinations_df = pd.DataFrame(columns=['Treatment Column', 'Treatment Concentration', 'Cell Column', 'Cell Concentration'])

    treatment_values = []
    treatment_indices = []
    culture_values = []
    culture_indices = []

    for i in range (1,13):
        antibiotic_concentration_type = ORIGINAL_ANTIBIOTIC_CONCENTRATION[i-1]
        if(antibiotic_concentration_type != 0 and antibiotic_concentration_type != None):
            for j in range(1,6):
                treatment_values.append(antibiotic_concentration_type/2**j)
                treatment_indices.append(i)
            treatment_values.append(0)
            treatment_indices.append(6)

    for i in range(1,13):
        cell_concentration_type = ORIGINAL_CELL_CONCENTRATION[i-1]/10
        if(cell_concentration_type != 0 and antibiotic_concentration_type != None):
            culture_values.append(cell_concentration_type)
            culture_indices.append(i)

    for x in range(0, len(treatment_values)):
        for y in range(0, len(culture_values)):
            latest_row = {
                'Treatment Column': treatment_indices[x], 
                'Treatment Concentration': treatment_values[x],
                'Cell Column': culture_indices[y],
                'Cell Concentration': culture_values[y] 
            }
            latest_row_df = pd.DataFrame(latest_row, index=[0])
            combinations_df = pd.concat([combinations_df, latest_row_df], ignore_index=True)

    return combinations_df
    
def delete_experiment_excel_file():
    os.remove(EXPERIMENT_FILE_PATH)

def determine_payload_from_excel():
    print("Run Log Starts Now")
    folder_path = str(pathlib.Path().resolve()) + "\\growth_app\\active_runs"
    #folder_path = str(pathlib.Path().resolve()) + "/active_runs"
    files = os.listdir(folder_path)
    excel_files = [file for file in files if file.endswith(".xlsx")]
    sorted_files = sorted(excel_files, key=lambda x: os.path.getmtime(os.path.join(folder_path, x)))
    path_name = os.path.join(folder_path, sorted_files[0])
    EXPERIMENT_FILE_PATH = path_name
    print(path_name)
    workbook = openpyxl.load_workbook(filename=path_name)
    worksheet = workbook['Complete_Run_Layout']
    experiment_iterations = worksheet['B1'].value
    incubation_time_hours = worksheet['E1'].value
    incubation_time_seconds = incubation_time_hours * 3600
    added_items = 0
    for i in range(2,13):
        column_letter = chr(ord('@')+i)
        run_number_cell_id = column_letter + "3"
        media_type_cell_id = column_letter + "4"
        culture_type_cell_id = column_letter + "5"
        if(worksheet[run_number_cell_id].value != None and worksheet[media_type_cell_id].value != None and worksheet[culture_type_cell_id].value != None):      
            MEDIA_PAYLOAD.append(worksheet[media_type_cell_id].value)
            CULTURE_PAYLOAD.append(worksheet[culture_type_cell_id].value) 
            added_items = added_items + 1
    if(len(MEDIA_PAYLOAD) != experiment_iterations):
        experiment_iterations = len(MEDIA_PAYLOAD)

    for i in range(0,12):
        original_concentration = ORIGINAL_ANTIBIOTIC_CONCENTRATION[i]/10
        single_plate_treatment_columns = []
        for iterations in range(0,2):
            for j in range (1,6):
                single_plate_treatment_columns.append(original_concentration/2**j)
            single_plate_treatment_columns.append(0)
        TOTAL_TREATMENT_COLUMN_CONCENTRATION.append(single_plate_treatment_columns)

    for i in range(0,12):
        original_concentration = ORIGINAL_CELL_CONCENTRATION[i]
        single_plate_cell_columns = []
        for iterations in range(0,12):
            single_plate_cell_columns.append(original_concentration/10)
        TOTAL_CELL_COLUMN_CONCENTRATION.append(single_plate_cell_columns)
    
    return experiment_iterations, incubation_time_seconds

def train_model():
    global TENSORFLOW_MODEL
    training_df = pd.DataFrame(columns=['Treatment Column', 'Treatment Concentration', 'Cell Column', 'Cell Concentration', 'Growth Rate'])
    #folder_path = str(pathlib.Path().resolve()) + "/completed_runs"
    folder_path = str(pathlib.Path().resolve()) + "\\bio_workcell\\completed_runs\\"
    path_name = folder_path + COMPLETED_FILE_NAME
    completed_workbook = openpyxl.load_workbook(path_name)
    for sheet_name in completed_workbook.sheetnames:
        current_sheet = completed_workbook[sheet_name]
        for i in range(2, 98):
            latest_row = {
                'Treatment Column': current_sheet["A" + str(int(i))].value, 
                'Treatment Concentration': current_sheet["B" + str(int(i))].value,
                'Cell Column': current_sheet["C" + str(int(i))].value,
                'Cell Concentration': current_sheet["D" + str(int(i))].value,
                'Growth Rate' : current_sheet["E" + str(int(i))].value,
            } 
            latest_row_df = pd.DataFrame(latest_row, index=[0])
            training_df = pd.concat([training_df, latest_row_df], ignore_index=True)

    numeric_columns = ['Treatment Column', 'Treatment Concentration', 'Cell Column', 'Cell Concentration', 'Growth Rate']
    training_df[numeric_columns] = training_df[numeric_columns].apply(pd.to_numeric, errors='coerce')

    training_df = training_df.dropna()  # Drop rows with missing or non-numeric values

    X = training_df[['Treatment Column', 'Treatment Concentration', 'Cell Column', 'Cell Concentration']]
    y = training_df['Growth Rate']

    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

    TENSORFLOW_MODEL.compile(optimizer='adam', loss='mse')
    TENSORFLOW_MODEL.fit(X_train, y_train, epochs=10, batch_size=32, validation_data=(X_test, y_test))

def process_results():
    results_path = str(pathlib.Path().resolve()) + "\\growth_app\\demo_data"
    files = os.listdir(results_path)
    excel_files = [file for file in files if file.endswith(".xlsx")]
    print(excel_files)
    print(HIDEX_UPLOADS)
    recent_files = sorted(excel_files, key=lambda x: os.path.getmtime(os.path.join(results_path, x)))[:len(HIDEX_UPLOADS)]
    print(recent_files)
    t0_run_ids = []
    t12_run_ids = []
    t0_excel_files = []
    t12_excel_files = []
    for i in range (0, len(HIDEX_UPLOADS)):
        upload_id = HIDEX_UPLOADS[i]
        excel_file_id = recent_files[i]
        if upload_id.startswith("T0_") :
            run_number = int(upload_id[3:])
            t0_run_ids.append(run_number)
            t0_excel_files.append(excel_file_id)
        elif upload_id.startswith("T12_"):
            run_number = int(upload_id[4:])
            t12_run_ids.append(run_number)
            t12_excel_files.append(excel_file_id)

    old_t0_run_ids = t0_run_ids
    t0_run_ids, t0_excel_files = zip(*sorted(zip(t0_run_ids, t0_excel_files)))
    t12_run_ids, t12_excel_files = zip(*sorted(zip(t12_run_ids, t12_excel_files)))

    if set(t0_run_ids) != set(t12_run_ids):
        mismatched_ids = set(t0_run_ids) ^ set(t12_run_ids)

        t0_run_ids, t0_excel_files = zip(*[(run_id, excel_file) for run_id, excel_file in zip(t0_run_ids, t0_excel_files) if run_id not in mismatched_ids])
        t12_run_ids, t12_excel_files = zip(*[(run_id, excel_file) for run_id, excel_file in zip(t12_run_ids, t12_excel_files) if run_id not in mismatched_ids])

    t0_run_ids = list(t0_run_ids)
    t0_excel_files = list(t0_excel_files)
    t12_run_ids = list(t12_run_ids)
    t12_excel_files = list(t12_excel_files)

    cell_columns = []
    antibiotic_string_columns = []
    barcodes = []

    for run_id in t0_run_ids:
        info_index = old_t0_run_ids.index(run_id)
        cell_columns.append(COMPLETED_CELL_COLUMNS[info_index])
        antibiotic_string_columns.append(COMPLETED_ANTIBIOTIC_COLUMNS[info_index])
        barcodes.append(PLATE_BARCODES[info_index])

    antibiotic_columns = []
    for string_column in antibiotic_string_columns:
        integer_value = int(string_column[3:])
        antibiotic_columns.append(integer_value)

    antibiotic_concentrations_list = []
    for antibiotic_column in antibiotic_columns:
        antibioitic_index = antibiotic_column - 1
        single_column_antibiotic_concentration_list = TOTAL_TREATMENT_COLUMN_CONCENTRATION[antibioitic_index]
        single_plate_all_antibiotic_concentrations = []
        for i in range(0,12):
            for j in range(0,8):
                single_plate_all_antibiotic_concentrations.append(single_column_antibiotic_concentration_list[i])
        antibiotic_concentrations_list.append(single_plate_all_antibiotic_concentrations)

    cell_concentrations_list = []
    for cell_column in cell_columns:
        cell_index = cell_column - 1
        single_column_cell_concentration_list = TOTAL_CELL_COLUMN_CONCENTRATION[cell_index]
        single_plate_all_cell_concentrations = []
        for i in range(0,12):
            for j in range(0,8):
                single_plate_all_cell_concentrations.append(single_column_cell_concentration_list[i])
        cell_concentrations_list.append(single_plate_all_antibiotic_concentrations)


    #completed_workbook =
    #folder_path = str(pathlib.Path().resolve()) + "/completed_runs"
    folder_path = str(pathlib.Path().resolve()) + "\\bio_workcell\\completed_runs\\"
    current_sheet_index = 1

    global CREATED_COMPLETED_FILE
    global COMPLETED_FILE_NAME

    if CREATED_COMPLETED_FILE:
        print("Completed File Name ", COMPLETED_FILE_NAME)
        path_name = folder_path + COMPLETED_FILE_NAME
        completed_workbook = openpyxl.load_workbook(file_name = path_name)
        num_sheets = len(completed_workbook.worksheets)
        current_sheet_index = num_sheets + 1

    else:
        print("Creating Excel Object")
        current_date = datetime.date.today()
        formatted_date = current_date.strftime("%m-%d-%Y")
        file_name = formatted_date + " Completed Run" + ".xlsx"
        COMPLETED_FILE_NAME = file_name
        completed_workbook = openpyxl.Workbook() 
        CREATED_COMPLETED_FILE = True
        os.makedirs(folder_path, exist_ok=True)
        completed_workbook.save(os.path.join(folder_path, COMPLETED_FILE_NAME))
        default_sheet = completed_workbook.active
        completed_workbook.remove(default_sheet)


    for i in range(0, len(t0_run_ids)):
        sheet_name = "Run " + str(int(current_sheet_index))
        current_sheet = completed_workbook.create_sheet(sheet_name)
        t0_path_name = os.path.join(results_path, t0_excel_files[i])
        t0_workbook = openpyxl.load_workbook(filename=t0_path_name)
        t0_worksheet = t0_workbook['Raw OD(590)']
        t12_path_name = os.path.join(results_path, t12_excel_files[i])
        t12_workbook = openpyxl.load_workbook(filename=t12_path_name)
        t12_worksheet = t12_workbook['Raw OD(590)']
        runs_df = pd.DataFrame(columns=['Treatment Column', 'Treatment Concentration', 'Cell Column', 'Cell Concentration', 'Growth Rate', 'T0 Reading', 'T12 Reading'])
        for j in range(10,106):
            hidex_data_index = "D" + str(int(j))
            t0_growth_value = t0_worksheet[hidex_data_index].value
            t12_growth_value = t12_worksheet[hidex_data_index].value
            growth_rate = t12_growth_value - t0_growth_value
            print("antibiotic concentrations ", antibiotic_concentrations_list)
            print("cell concdentrations ", cell_concentrations_list)
            latest_row = {
                'Treatment Column': antibiotic_columns[i], 
                'Treatment Concentration': antibiotic_concentrations_list[i][j-10],
                'Cell Column': cell_columns[i],
                'Cell Concentration': cell_concentrations_list[i][j-10],
                'Growth Rate' : growth_rate,
                'T0 Reading' : t0_growth_value,
                'T12 Reading' : t12_growth_value
            }

            latest_row_df = pd.DataFrame(latest_row, index=[0])
            runs_df = pd.concat([runs_df, latest_row_df], ignore_index=True)
    
            for row in dataframe_to_rows(runs_df, index=False, header=True):
                current_sheet.append(row)

            current_sheet['I1'] = "Barcode Number"
            current_sheet['J1'] = barcodes[i]

    completed_workbook.save(folder_path + COMPLETED_FILE_NAME)

def save_model():
    TENSORFLOW_MODEL.save(AI_MODEL_FILE_PATH)

def assign_barcode():
    current_barcode = return_barcode()
    PLATE_BARCODES.append(current_barcode)

def return_barcode():
    barcode = ""
    for _ in range(13):
        digit = random.randint(0, 9)
        barcode += str(digit)
    return barcode

def run_experiment(total_iterations, incubation_time_sec): 
    iterations = 0
    removals = 0
    incubation_start_times = []
    print("Total Experimental Runs ", total_iterations)
    print("Current Iteration Variable: ", iterations)
    print("Total Iterations: ", incubation_time_sec)
    # The experiment will run until all of the plates are used (indicated by iterations < total_iterations) and there are no more plates in the incubator (indicated by len(incubation_start_times) != 0)
    while(iterations < total_iterations or len(incubation_start_times) != 0):
        #Check to see if there are any more plates to run, indicated by total_iterations
        if(iterations < total_iterations):
            #Debug Log
            print("Starting Experiment ", iterations, ": Started Loop")
            #Set up the experiment based on the number of iterations passed.
            setup(iterations)
            #Calculate the ID of the plate needed for incubation based on the number of iterations that have passed
            liconic_id = iterations + 1
            #Run the experiment from the Hudson Solo step to the incubation step at a specified Liconic ID
            print("Starting T0 Reading")
            T0_Reading(liconic_id)
            print("Finished T0 Reading")
            assign_barcode()
            #Add the time of the incubation start to the array of 96 well plates that are currently in the incubator
            incubation_start_times.append(round(time.time()))
            #Since an iteration has now passed (the plate is in the incubator), increase the index of incubation iterations
            iterations = iterations + 1
            #Debug Log
            print("Completed Iterations: ", iterations, " ... Start Times: " , incubation_start_times)
            #Based on the total number of completed incubation iterations, determine what needs to be disposed of from the experimental setup.
            if(iterations % 2 == 0):
                print("Starting Disposal")
                dispose(iterations)
                print("Ending Disposal")
        #Check to see if delta current time and the time at which the well plate currently incubating longest exceeds the incubation time.
        if(round(time.time()) - incubation_start_times[0] > incubation_time_sec):
            #Debug Log
            print("Finishing Up Experiment ", removals, ": Ending Loop")
            #Calcuate the ID of the 96 well plate needed for removal from the incubator based on the number of plates that have already been removed.
            liconic_id = removals + 1
            #Complete the experiment starting from the removal of the 96 well plate at the specified incubation ID and ending at a Hidex T12 Reading.
            print("Starting T12 Reading")
            T12_Reading(liconic_id)
            print("Ending T12 Reading")
            #Remove the incubation start time of the plate that was just read from the array to update the new longest incubating well plate
            incubation_start_times.pop(0)
            #Increase the total number of removals that have now occurred.
            removals = removals + 1

def dispose(completed_iterations):
    #Set the default disposal index of the serial dilution plate to Stack 2.
    disposal_index = "Stack2"
    #To identify the LidNest location of where certain serial dilution plates must be held, divide the completed iterations by 2
    stack_type = completed_iterations/2
    #For the first two serial dilution plates, define the disposal index as LidNest 1 for the first serial dilution plate and LidNest 2 for the second serial dilution plate
    if(stack_type <= 2):
        disposal_index = "LidNest" + str(int(stack_type))
    #For the fourth serial dilution plate, define the disposal index as LidNest 3.
    #We are not defining the disposal index as LidNest 3 for the third serial dilution plate because there will already be a growth media plate on LidNest 3. This growth media plate will be removed in the subsequent setup function
    if(stack_type == 4):
        disposal_index = "LidNest3"
    #Add the disposal index to the payload
    payload={
        'disposal_location':  disposal_index,    
        }
    #Run the despose Yaml File with the dedicated disposal location
    run_WEI(DISPOSE_BOX_PLATE_FILE_PATH, payload, False)
    if(stack_type % 3 == 0):
    #If the Stack Type is a multiple of 3 (6 complete iterations of the Hudson experiment have occurred), dispose of the growth media deep well plate
        run_WEI(DISPOSE_GROWTH_MEDIA_FILE_PATH, None, False)

def setup(iteration_number):
    #If currently on an even number of iterations, add a tip box, serial dilution plate, and 96 well plate to the experiment.
    if(iteration_number % 2 == 0):
        #Identify the Tip Box Position Index, so it can be refilled on the Hudson Client
        payload={
                'tip_box_position': "3",
            }
        #Run the Yaml file that outlines the setup procedure for the tip box, serial dilution plate, and 96 well plate.
        print("Starting Complete Hudson Setup")
        run_WEI(COMPLETE_HUDSON_SETUP_FILE_PATH, payload, False)
        print("Finished Complete Setup")
        #If currently on a number of iterations that is a factor of 6, add a growth media plate to the experiment as well
        if(iteration_number % 6 == 0):
            #Specify the LidNest index of the growth well media plate that will be added to the setup (This equation assumes that there are only two growth media plates on LidNest 2 and LidNest 3 respectively for a total of 12 runs)
            LidNest_index = 2 + iteration_number/6
            #Convert the index to a readable string
            plateCrane_readable_index = "LidNest" + str(int(LidNest_index))
            print("LidNest Being Used: ", plateCrane_readable_index)
            #Add the LidNest Index to the payload
            payload={
                'lidnest_index':  plateCrane_readable_index,
            }
            #Run the Yaml file that outlines the setup procedure for the growth media plate
            print("Starting Growth Media Setup")
            run_WEI(SETUP_GROWTH_MEDIA_FILE_PATH, payload, False)
            print("Finished Growth Media Setup!")
    #If currently on an even number of iterations, add a 96 well plate to the experiment.
    else: 
        #Run the Yaml file that outlines the setup procedure for ONLY the 96 well plate
        run_WEI(STREAMLINED_HUDSON_SETUP_FILE_PATH, None, False)
        
def refreshHidex():
    run_WEI(HIDEX_OPEN_CLOSE_FILE_PATH, None, False)

def T0_Reading(liconic_plate_id):
    plate_id = '' + str(int(liconic_plate_id))
    treatment_col_id = "col" + str(int(MEDIA_PAYLOAD[liconic_plate_id-1]))
    culture_col_id = int(CULTURE_PAYLOAD[liconic_plate_id-1])
    payload={
        'temp': 37.0, 
        'humidity': 95.0,
        'shaker_speed': 30,
        "stacker": 1, 
        "slot": 1,
        "treatment": treatment_col_id, # string of treatment name. Ex. "col1", "col2"
        "culture_column": culture_col_id,  # int of cell culture column. Ex. 1, 2, 3, etc.
        "culture_dil_column": 1, # int of dilution column for 1:10 culture dilutions. Ex. 1, 2, 3, etc.
        "media_start_column": 1,  # int of column to draw media from (requires 2 columns, 1 means columns 1 and 2) Ex. 1, 3, 5, etc.
        "treatment_dil_half": 1,  #  int of which plate half to use for treatment serial dilutions. Options are 1 or 2. 
        "incubation_plate_id" : plate_id,        
        }

    # from somewhere import create_hso? or directly the solo script
    # hso_1, hso_1_lines, hso_1_basename = package_hso(solo_multi_step1.generate_hso_file, payload, "/home/rpl/wei_temp/solo_temp1.hso") 
    # hso_2, hso_2_lines, hso_2_basename = package_hso(solo_multi_step2.generate_hso_file, payload, "/home/rpl/wei_temp/solo_temp2.hso")  
    # hso_3, hso_3_lines, hso_3_basename = package_hso(solo_multi_step3.generate_hso_file, payload, "/home/rpl/wei_temp/solo_temp3.hso")  

    # # update payload with solo hso details
    # payload['hso_1'] = hso_1
    # payload['hso_1_lines'] = hso_1_lines
    # payload['hso_1_basename'] = hso_1_basename

    # payload['hso_2'] = hso_2
    # payload['hso_2_lines'] = hso_2_lines
    # payload['hso_2_basename'] = hso_2_basename

    # payload['hso_3'] = hso_3
    # payload['hso_3_lines'] = hso_3_lines
    # payload['hso_3_basename'] = hso_3_basename

    # #run Growth Create Plate
    # run_WEI(CREATE_PLATE_T0_FILE_PATH, payload, True)

    hidex_upload_id = "T0_" + str(int(liconic_plate_id))

    HIDEX_UPLOADS.append(hidex_upload_id)
    COMPLETED_ANTIBIOTIC_COLUMNS.append(treatment_col_id)
    COMPLETED_CELL_COLUMNS.append(culture_col_id)

def T12_Reading(liconic_plate_id):
    plate_id = '' + str(int(liconic_plate_id))

    payload={
        'temp': 37.0, 
        'humidity': 95.0,
        'shaker_speed': 30,
        "stacker": 1, 
        "slot": 1,
        "treatment": "col1", # string of treatment name. Ex. "col1", "col2"
        "culture_column": 1,  # int of cell culture column. Ex. 1, 2, 3, etc.
        "culture_dil_column": 1, # int of dilution column for 1:10 culture dilutions. Ex. 1, 2, 3, etc.
        "media_start_column": 1,  # int of column to draw media from (requires 2 columns, 1 means columns 1 and 2) Ex. 1, 3, 5, etc.
        "treatment_dil_half": 1,  #  int of which plate half to use for treatment serial dilutions. Options are 1 or 2.
        "incubation_plate_id" : plate_id,
        }

    # from somewhere import create_hso? or directly the solo script
    # hso_1, hso_1_lines, hso_1_basename = package_hso(solo_multi_step1.generate_hso_file, payload, "/home/rpl/wei_temp/solo_temp1.hso") 
    # hso_2, hso_2_lines, hso_2_basename = package_hso(solo_multi_step2.generate_hso_file, payload, "/home/rpl/wei_temp/solo_temp2.hso")  
    # hso_3, hso_3_lines, hso_3_basename = package_hso(solo_multi_step3.generate_hso_file, payload, "/home/rpl/wei_temp/solo_temp3.hso")  

    # # update payload with solo hso details
    # payload['hso_1'] = hso_1
    # payload['hso_1_lines'] = hso_1_lines
    # payload['hso_1_basename'] = hso_1_basename

    # payload['hso_2'] = hso_2
    # payload['hso_2_lines'] = hso_2_lines
    # payload['hso_2_basename'] = hso_2_basename

    # payload['hso_3'] = hso_3
    # payload['hso_3_lines'] = hso_3_lines
    # payload['hso_3_basename'] = hso_3_basename

    # # #run Growth Create Plate
    # run_WEI(READ_PLATE_T12_FILE_PATH, payload, True)
    
    hidex_upload_id = "T12_" + str(int(liconic_plate_id))
    HIDEX_UPLOADS.append(hidex_upload_id)

def run_WEI(file_location, payload_class, Hidex_Used):
    print(file_location)
    # flow_info = exp.run_job(Path(file_location).resolve(), payload=payload_class, simulate=False)

    # flow_status = exp.query_job(flow_info["job_id"])
    # while(flow_status["status"] != "finished" and flow_status["status"] != "failure"):
    #     flow_status = exp.query_job(flow_info["job_id"])
    #     time.sleep(3)

    # run_info = flow_status["result"]
    # run_info["run_dir"] = Path(run_info["run_dir"])
    # print(run_info)

    # if Hidex_Used:
    #     print("Starting Up Hidex")
    #     hidex_file_path = run_info["hist"]["run Hidex"]["action_msg"]
    #     hidex_file_path = hidex_file_path.replace('\\', '/')
    #     hidex_file_path = hidex_file_path.replace("C:/", "/C/")
    #     flow_title = Path(hidex_file_path) #Path(run_info["hist"]["run_assay"]["step_response"])
    #     fname = flow_title.name
    #     flow_title = flow_title.parents[0]

    #     c2_flow("hidex_test", str(fname.split('.')[0]), hidex_file_path, flow_title, fname, exp)
    #     print("Finished Uplodaing to Globus")

if __name__ == "__main__":
    main()


#!/usr/bin/env python3



# Previous Functions/Workflows
# HIDEX_IDLE_THRESHOLD_SECONDS = 3600

# def two():
#     iterations = 0
#     removals = 0
#     incubation_start_times = []

#     hidex_refresh_time = round(time.time())
#     EXPERIMENT_ITERATIONS = 2
#     while(iterations < EXPERIMENT_ITERATIONS or len(incubation_start_times) == 0):
#         if(iterations < EXPERIMENT_ITERATIONS):
#             setup(iterations)
#             liconic_id = iterations + 1
#             T0_Reading(liconic_id)
#             incubation_start_times.append(round(time.time()))
#             iterations = iterations + 1
#             if(iterations % 2 == 0):
#                 dispose()
#             #hidex_refresh_time = round(time.time())
#         if(round(time.time()) - incubation_start_times[0] > INCUBATION_TIME_SECONDS):
#             liconic_id = removals + 1
#             T12_Reading(liconic_id)
#             incubation_start_times.pop(0)
#             removals = removals + 1
#             #hidex_refresh_time = round(time.time())
#         #if(round(time.time()) - hidex_refresh_time < (HIDEX_IDLE_THRESHOLD_SECONDS - 20*60)):
#             #refreshHidex()

# def six():
#     iterations = 0
#     removals = 0
#     incubation_start_times = []

#     EXPERIMENT_ITERATIONS = 6
#     hidex_refresh_time = round(time.time())
#     EXPERIMENT_ITERATIONS = 2
#     while(iterations < EXPERIMENT_ITERATIONS or len(incubation_start_times) == 0):
#         if(iterations < EXPERIMENT_ITERATIONS):
#             setup(iterations)
#             liconic_id = iterations + 1
#             T0_Reading(liconic_id)
#             incubation_start_times.append(round(time.time()))
#             iterations = iterations + 1
#             if(iterations % 2 == 0):
#                 dispose()
#             #hidex_refresh_time = round(time.time())
#         if(round(time.time()) - incubation_start_times[0] > INCUBATION_TIME_SECONDS):
#             liconic_id = removals + 1
#             T12_Reading(liconic_id)
#             incubation_start_times.pop(0)
#             removals = removals + 1
#             #hidex_refresh_time = round(time.time())
#         #if(round(time.time()) - hidex_refresh_time < (HIDEX_IDLE_THRESHOLD_SECONDS - 20*60)):
#             #refreshHidex()

# def twelve():
#     iterations = 0
#     removals = 0
#     incubation_start_times = []

#     while(iterations < EXPERIMENT_ITERATIONS or len(incubation_start_times) == 0):
#         if(iterations < EXPERIMENT_ITERATIONS):
#             setup(iterations)
#             liconic_id = iterations + 1
#             T0_Reading(liconic_id)
#             incubation_start_times.append(round(time.time()))
#             iterations = iterations + 1
#             if(iterations % 2 == 0):
#                 dispose(iterations)
#             #hidex_refresh_time = round(time.time())
#         if(round(time.time()) - incubation_start_times[0] > INCUBATION_TIME_SECONDS):
#             liconic_id = removals + 1
#             T12_Reading(liconic_id)
#             incubation_start_times.pop(0)
#             removals = removals + 1
#             #hidex_refresh_time = round(time.time())
#         #if(round(time.time()) - hidex_refresh_time < (HIDEX_IDLE_THRESHOLD_SECONDS - 20*60)):
#             #refreshHidex()

# def base():
#     iterations = 0
#     removals = 0
#     incubation_start_times = []
#     while(iterations < EXPERIMENT_ITERATIONS or len(incubation_start_times) == 0):
#         if(iterations < EXPERIMENT_ITERATIONS):   
#             T0_Reading()
#             incubation_start_times.append(round(time.time()))
#             iterations = iterations + 1
#         if(round(time.time()) - incubation_start_times[0] > INCUBATION_TIME_SECONDS):
#             T12_Reading()
#             incubation_start_times.pop(0)

